# -*- coding: utf-8 -*-
from qgis.PyQt.QtWidgets import QAction, QMessageBox
from qgis.PyQt.QtGui import QIcon
from qgis.core import QgsProject, QgsVectorLayer
from .xlsx_furigana_delete_dialog import XlsxFuriganaDeleteDialog

from openpyxl import load_workbook
import os
# shutilモジュールを追加
import shutil

class XlsxFuriganaDelete:

    def __init__(self, iface):
        self.iface = iface
        self.action = None
        self.dlg = None

    def initGui(self):
        icon_path = os.path.join(os.path.dirname(__file__), "icon.png")

        self.action = QAction(
            QIcon(icon_path),
            "ふりがな削除（Excel）",
            self.iface.mainWindow()
        )
        self.action.triggered.connect(self.run)

        self.iface.addToolBarIcon(self.action)
        self.iface.addPluginToMenu("xlsx_furigana_delete", self.action)

    def unload(self):
        self.iface.removeToolBarIcon(self.action)
        self.iface.removePluginMenu("xlsx_furigana_delete", self.action)

    def run(self):
        self.dlg = XlsxFuriganaDeleteDialog()

        layer = self.iface.activeLayer()
        if layer:
            self.dlg.layer_combo.setCurrentText(layer.name())

        self.dlg.buttonBox.accepted.connect(self.process)
        self.dlg.buttonBox.rejected.connect(self.dlg.close)
        self.dlg.show()

    def process(self):
        layer_name = self.dlg.layer_combo.currentText()
        layers = QgsProject.instance().mapLayersByName(layer_name)
        if not layers:
            QMessageBox.warning(None, "エラー", "レイヤが見つかりません。")
            # 処理完了（失敗）後にダイアログを閉じる
            self.dlg.close() 
            return

        layer = layers[0]
        source = layer.source()
        file_path = source.split("|")[0]

        if not file_path.lower().endswith(".xlsx"):
            QMessageBox.warning(None, "エラー", "このレイヤは Excel（xlsx）ファイルではありません。")
            # 処理完了（失敗）後にダイアログを閉じる
            self.dlg.close()
            return

        # 1. バックアップファイルのパスを生成
        dir_name = os.path.dirname(file_path)
        base_name = os.path.basename(file_path)
        name, ext = os.path.splitext(base_name)
        backup_file_path = os.path.join(dir_name, f"{name}_Backup{ext}")

        # 2. 対象ファイルをバックアップとしてコピー
        try:
            shutil.copy2(file_path, backup_file_path) # copy2はメタデータもコピーします
            QMessageBox.information(None, "バックアップ", f"バックアップを作成しました:\n{backup_file_path}")
        except Exception as e:
            QMessageBox.critical(None, "エラー", f"バックアップ作成中にエラー:\n{str(e)}")
            # 処理完了（失敗）後にダイアログを閉じる
            self.dlg.close()
            return


        # --- Excelのふりがな削除 ---
        try:
            wb = load_workbook(file_path)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if hasattr(cell, "phonetic"):
                            cell.phonetic = None

            wb.save(file_path)

        except Exception as e:
            QMessageBox.critical(None, "エラー", f"Excel処理中にエラー:\n{str(e)}")
            # 処理完了（失敗）後にダイアログを閉じる
            self.dlg.close()
            return

        # --- レイヤを再ロードしてフィールド名更新 ---
        project = QgsProject.instance()
        layer_id = layer.id()
        layer_name = layer.name()

        # 1. 現在のレイヤを削除
        project.removeMapLayer(layer_id)

        # 2. 同じファイルパスで再読み込み
        new_layer = QgsVectorLayer(file_path, layer_name, "ogr")
        if not new_layer.isValid():
            QMessageBox.critical(None, "エラー", "Excel レイヤの再読み込みに失敗しました。")
            # 処理完了（失敗）後にダイアログを閉じる
            self.dlg.close()
            return

        project.addMapLayer(new_layer)

        QMessageBox.information(None, "完了", "ふりがな削除し、レイヤとフィールド名を更新しました。")
        
        # 3. 処理完了後にダイアログを閉じる
        self.dlg.close()